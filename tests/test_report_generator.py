"""
Unit tests for report_generator module.
"""
import json
import os
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest
import pandas as pd
from openpyxl import load_workbook

from src.report_generator import ReportGenerator, create_sample_report
from src.models import CortinillaResult, AccumulatedResults, Occurrence


class TestReportGenerator:
    """Test cases for ReportGenerator class."""
    
    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for testing."""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield temp_dir
    
    @pytest.fixture
    def report_generator(self, temp_dir):
        """Create ReportGenerator instance with temporary directory."""
        return ReportGenerator(temp_dir)
    
    @pytest.fixture
    def sample_cortinilla_result(self):
        """Create sample CortinillaResult for testing."""
        occurrences = {
            "buenos días": [
                Occurrence(start_time=120.5, end_time=122.3, confidence=0.95, text="buenos días"),
                Occurrence(start_time=1800.2, end_time=1802.1, confidence=0.92, text="buenos días")
            ],
            "buenas tardes": [
                Occurrence(start_time=900.1, end_time=901.8, confidence=0.88, text="buenas tardes")
            ]
        }
        
        cortinillas_by_type = {
            "buenos días": 2,
            "buenas tardes": 1
        }
        
        return CortinillaResult(
            channel="TestChannel",
            timestamp=datetime(2024, 1, 15, 14, 0, 0),
            audio_duration=3600.0,
            total_cortinillas=3,
            cortinillas_by_type=cortinillas_by_type,
            cortinillas_details=occurrences,
            overlap_filtered=False,
            overlap_duration=None
        )
    
    def test_init_creates_data_directory(self, temp_dir):
        """Test that ReportGenerator creates data directory if it doesn't exist."""
        data_dir = Path(temp_dir) / "test_data"
        assert not data_dir.exists()
        
        generator = ReportGenerator(str(data_dir))
        assert data_dir.exists()
        assert data_dir.is_dir()
    
    def test_update_json_report_creates_new_file(self, report_generator, sample_cortinilla_result):
        """Test creating new JSON report file."""
        report_generator.update_json_report(sample_cortinilla_result)
        
        json_path = Path(report_generator.data_dir) / "TestChannel_results.json"
        assert json_path.exists()
        
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        assert data["channel"] == "TestChannel"
        assert data["total_hours_processed"] == 1
        assert len(data["results"]) == 1
        assert data["results"][0]["total_cortinillas"] == 3
    
    def test_update_json_report_appends_to_existing(self, report_generator, sample_cortinilla_result):
        """Test appending to existing JSON report file."""
        # Create first result
        report_generator.update_json_report(sample_cortinilla_result)
        
        # Create second result
        second_result = CortinillaResult(
            channel="TestChannel",
            timestamp=datetime(2024, 1, 15, 15, 0, 0),
            audio_duration=3600.0,
            total_cortinillas=2,
            cortinillas_by_type={"buenos días": 1, "buenas noches": 1},
            cortinillas_details={
                "buenos días": [Occurrence(120.5, 122.3, 0.95, "buenos días")],
                "buenas noches": [Occurrence(3500.1, 3502.0, 0.90, "buenas noches")]
            },
            overlap_filtered=True,
            overlap_duration=30.0
        )
        
        report_generator.update_json_report(second_result)
        
        json_path = Path(report_generator.data_dir) / "TestChannel_results.json"
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        assert data["total_hours_processed"] == 2
        assert len(data["results"]) == 2
        assert data["results"][1]["overlap_filtered"] is True
        assert data["results"][1]["overlap_duration"] == 30.0
    
    def test_update_excel_report_creates_workbook(self, report_generator, sample_cortinilla_result):
        """Test creating Excel report with multiple sheets."""
        report_generator.update_excel_report(sample_cortinilla_result)
        
        excel_path = Path(report_generator.data_dir) / "TestChannel_results.xlsx"
        assert excel_path.exists()
        
        # Load workbook and check sheets
        wb = load_workbook(excel_path)
        expected_sheets = ["Resumen", "Detalles por Hora", "Desglose Cortinillas"]
        assert all(sheet in wb.sheetnames for sheet in expected_sheets)
        
        # Check summary sheet content
        summary_ws = wb["Resumen"]
        assert summary_ws["A1"].value == "Métrica"
        assert summary_ws["B1"].value == "Valor"
        assert summary_ws["B2"].value == "TestChannel"  # Channel name
        assert summary_ws["B3"].value == 1  # Total hours processed
    
    def test_load_existing_results_returns_none_for_nonexistent(self, report_generator):
        """Test loading results when file doesn't exist."""
        result = report_generator.load_existing_results("NonExistentChannel")
        assert result is None
    
    def test_load_existing_results_loads_valid_data(self, report_generator, sample_cortinilla_result):
        """Test loading existing results from JSON file."""
        # First create a report
        report_generator.update_json_report(sample_cortinilla_result)
        
        # Then load it
        accumulated = report_generator.load_existing_results("TestChannel")
        
        assert accumulated is not None
        assert accumulated.channel == "TestChannel"
        assert accumulated.total_hours_processed == 1
        assert len(accumulated.results) == 1
        assert accumulated.results[0].total_cortinillas == 3
    
    def test_get_channel_summary_empty_channel(self, report_generator):
        """Test getting summary for channel with no data."""
        summary = report_generator.get_channel_summary("EmptyChannel")
        
        expected = {
            "channel": "EmptyChannel",
            "total_hours": 0,
            "total_cortinillas": 0,
            "cortinillas_by_type": {},
            "last_processed": None
        }
        
        assert summary == expected
    
    def test_get_channel_summary_with_data(self, report_generator, sample_cortinilla_result):
        """Test getting summary for channel with data."""
        report_generator.update_json_report(sample_cortinilla_result)
        
        summary = report_generator.get_channel_summary("TestChannel")
        
        assert summary["channel"] == "TestChannel"
        assert summary["total_hours"] == 1
        assert summary["total_cortinillas"] == 3
        assert summary["cortinillas_by_type"]["buenos días"] == 2
        assert summary["cortinillas_by_type"]["buenas tardes"] == 1
        assert summary["last_processed"] is not None
    
    def test_cortinilla_result_serialization_roundtrip(self, report_generator, sample_cortinilla_result):
        """Test that CortinillaResult can be serialized and deserialized correctly."""
        # Convert to dict and back
        result_dict = report_generator._cortinilla_result_to_dict(sample_cortinilla_result)
        restored_result = report_generator._dict_to_cortinilla_result(result_dict)
        
        assert restored_result.channel == sample_cortinilla_result.channel
        assert restored_result.timestamp == sample_cortinilla_result.timestamp
        assert restored_result.audio_duration == sample_cortinilla_result.audio_duration
        assert restored_result.total_cortinillas == sample_cortinilla_result.total_cortinillas
        assert restored_result.cortinillas_by_type == sample_cortinilla_result.cortinillas_by_type
        assert restored_result.overlap_filtered == sample_cortinilla_result.overlap_filtered
        assert restored_result.overlap_duration == sample_cortinilla_result.overlap_duration
        
        # Check cortinillas_details
        for cortinilla_type, occurrences in restored_result.cortinillas_details.items():
            original_occurrences = sample_cortinilla_result.cortinillas_details[cortinilla_type]
            assert len(occurrences) == len(original_occurrences)
            for i, occ in enumerate(occurrences):
                orig_occ = original_occurrences[i]
                assert occ.start_time == orig_occ.start_time
                assert occ.end_time == orig_occ.end_time
                assert occ.confidence == orig_occ.confidence
                assert occ.text == orig_occ.text
    
    def test_excel_formatting_applied(self, report_generator, sample_cortinilla_result):
        """Test that Excel formatting is properly applied."""
        report_generator.update_excel_report(sample_cortinilla_result)
        
        excel_path = Path(report_generator.data_dir) / "TestChannel_results.xlsx"
        wb = load_workbook(excel_path)
        
        # Check header formatting in summary sheet
        summary_ws = wb["Resumen"]
        header_cell = summary_ws["A1"]
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb == "00366092"
        
        # Check details sheet has proper headers
        details_ws = wb["Detalles por Hora"]
        expected_headers = [
            "Fecha y Hora", "Duración Audio (min)", "Total Cortinillas",
            "Filtrado por Solapamiento", "Duración Solapamiento (min)"
        ]
        
        for i, expected_header in enumerate(expected_headers, 1):
            assert details_ws.cell(row=1, column=i).value == expected_header
    
    def test_multiple_channels_separate_files(self, report_generator):
        """Test that different channels create separate files."""
        # Create results for two different channels
        result1 = CortinillaResult(
            channel="Channel1",
            timestamp=datetime.now(),
            audio_duration=3600.0,
            total_cortinillas=2,
            cortinillas_by_type={"buenos días": 2},
            cortinillas_details={"buenos días": [Occurrence(120.5, 122.3, 0.95, "buenos días")]},
            overlap_filtered=False,
            overlap_duration=None
        )
        
        result2 = CortinillaResult(
            channel="Channel2",
            timestamp=datetime.now(),
            audio_duration=3600.0,
            total_cortinillas=1,
            cortinillas_by_type={"buenas tardes": 1},
            cortinillas_details={"buenas tardes": [Occurrence(900.1, 901.8, 0.88, "buenas tardes")]},
            overlap_filtered=False,
            overlap_duration=None
        )
        
        report_generator.update_json_report(result1)
        report_generator.update_json_report(result2)
        
        # Check that separate files were created
        json_path1 = Path(report_generator.data_dir) / "Channel1_results.json"
        json_path2 = Path(report_generator.data_dir) / "Channel2_results.json"
        
        assert json_path1.exists()
        assert json_path2.exists()
        
        # Verify content is separate
        with open(json_path1, 'r', encoding='utf-8') as f:
            data1 = json.load(f)
        with open(json_path2, 'r', encoding='utf-8') as f:
            data2 = json.load(f)
        
        assert data1["channel"] == "Channel1"
        assert data2["channel"] == "Channel2"
        
        # Check the actual keys in cortinillas_by_type
        cortinillas1 = data1["results"][0]["cortinillas_by_type"]
        cortinillas2 = data2["results"][0]["cortinillas_by_type"]
        
        assert "buenos días" in cortinillas1
        assert cortinillas1["buenos días"] == 2
        assert "buenas tardes" in cortinillas2
        assert cortinillas2["buenas tardes"] == 1
    
    @patch('src.report_generator.logger')
    @patch('builtins.open', side_effect=PermissionError("Permission denied"))
    def test_error_handling_json_update(self, mock_open, mock_logger, report_generator, sample_cortinilla_result):
        """Test error handling during JSON update."""
        with pytest.raises(PermissionError):
            report_generator.update_json_report(sample_cortinilla_result)
        
        # Verify error was logged
        mock_logger.error.assert_called()
        error_call = mock_logger.error.call_args[0][0]
        assert "Error updating JSON report" in error_call
    
    @patch('src.report_generator.logger')
    @patch('src.report_generator.ReportGenerator.load_existing_results', side_effect=IOError("Disk full"))
    def test_error_handling_excel_update(self, mock_load, mock_logger, report_generator, sample_cortinilla_result):
        """Test error handling during Excel update."""
        with pytest.raises(IOError):
            report_generator.update_excel_report(sample_cortinilla_result)
        
        # Verify error was logged
        mock_logger.error.assert_called()
        error_call = mock_logger.error.call_args[0][0]
        assert "Error updating Excel report" in error_call


class TestCreateSampleReport:
    """Test cases for create_sample_report function."""
    
    def test_create_sample_report(self, tmp_path):
        """Test creating sample reports."""
        data_dir = str(tmp_path)
        
        create_sample_report("TestChannel", data_dir)
        
        # Check that files were created
        json_path = tmp_path / "TestChannel_results.json"
        excel_path = tmp_path / "TestChannel_results.xlsx"
        
        assert json_path.exists()
        assert excel_path.exists()
        
        # Verify JSON content
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        assert data["channel"] == "TestChannel"
        assert data["total_hours_processed"] == 3
        assert len(data["results"]) == 3
        
        # Verify Excel content
        wb = load_workbook(excel_path)
        assert "Resumen" in wb.sheetnames
        assert "Detalles por Hora" in wb.sheetnames
        assert "Desglose Cortinillas" in wb.sheetnames


if __name__ == "__main__":
    pytest.main([__file__])