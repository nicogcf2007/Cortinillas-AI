"""
Report generation and data persistence for Cortinillas AI.
Handles JSON and Excel report management with accumulative data storage.
"""
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from .models import CortinillaResult, CortinillaDetectionResult, AccumulatedResults, Occurrence, ProcessingExecution
    from .exceptions import ReportGenerationError, FileOperationError
    from .error_handler import ErrorHandler, create_error_context, safe_execute
except ImportError:
    from models import CortinillaResult, CortinillaDetectionResult, AccumulatedResults, Occurrence, ProcessingExecution
    from exceptions import ReportGenerationError, FileOperationError
    from error_handler import ErrorHandler, create_error_context, safe_execute


logger = logging.getLogger(__name__)


class ReportGenerator:
    """Handles report generation and data persistence for Cortinillas AI."""
    
    def __init__(self, data_dir: str = "data"):
        """
        Initialize the report generator.
        
        Args:
            data_dir: Directory where reports will be stored
        """
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(exist_ok=True)
        self.error_handler = ErrorHandler(max_retries=2, base_delay=1.0)
        self.logs_and_errors = []  # Store logs and errors for reporting
        logger.info(f"ReportGenerator initialized with data dir: {data_dir}")
    
    def _format_datetime(self, datetime_str: str) -> str:
        """
        Format datetime string to a more readable format without timezone.
        
        Args:
            datetime_str: ISO format datetime string
            
        Returns:
            Formatted datetime string
        """
        try:
            # Parse ISO format datetime (with or without timezone)
            if 'T' in datetime_str:
                if '+' in datetime_str or datetime_str.endswith('Z'):
                    # Has timezone info
                    dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
                else:
                    # No timezone info
                    dt = datetime.fromisoformat(datetime_str)
            else:
                # Already formatted or different format
                return datetime_str
            
            # Format as: "17 Sep 2025, 14:00"
            return dt.strftime("%d %b %Y, %H:%M")
        except Exception:
            # If parsing fails, return original string
            return datetime_str
    
    def add_log_entry(self, level: str, message: str, context: str = None):
        """
        Add a log entry for inclusion in reports.
        
        Args:
            level: Log level (INFO, WARNING, ERROR, CRITICAL)
            message: Log message
            context: Optional context information
        """
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "level": level,
            "message": message,
            "context": context or "N/A"
        }
        self.logs_and_errors.append(log_entry)
        
        # Keep only last 100 entries to prevent memory issues
        if len(self.logs_and_errors) > 100:
            self.logs_and_errors = self.logs_and_errors[-100:]
        
    def update_json_report(self, result: CortinillaDetectionResult) -> None:
        """
        Update the JSON report file with new cortinilla results.
        
        Args:
            result: CortinillaResult to add to the report
        """
        context = create_error_context(
            "update_json_report",
            channel=result.channel,
            timestamp=result.timestamp
        )
        
        try:
            json_path = self.data_dir / f"{result.channel.lower()}_results.json"
            
            # Load existing results or create new
            accumulated = safe_execute(
                self.load_existing_results,
                result.channel,
                default_return=None,
                error_handler=self.error_handler,
                context=f"{context} | load_existing"
            )
            
            if accumulated is None:
                accumulated = AccumulatedResults(
                    channel_name=result.channel,
                    total_executions=0,
                    total_cortinillas_found=0,
                    last_execution=None,
                    executions=[]
                )
            
            # Add cortinilla detection log
            if result.total_cortinillas > 0:
                cortinilla_summary = ", ".join([f"{phrase}: {len(occs)}" for phrase, occs in result.cortinillas_details.items() if occs])
                self.add_log_entry("INFO", f"Cortinillas detected: {cortinilla_summary}", f"Channel: {result.channel}")
            else:
                self.add_log_entry("INFO", "No cortinillas detected", f"Channel: {result.channel}")
            
            # Add overlap detection log if applicable
            if result.overlap_filtered:
                self.add_log_entry("INFO", f"Overlap detected and filtered: {result.overlap_duration:.2f}s removed", f"Channel: {result.channel}")
            
            # Create ProcessingExecution from CortinillaDetectionResult
            execution = ProcessingExecution(
                timestamp=result.timestamp.isoformat(),
                time_range=f"{result.start_time.strftime('%H:%M')} - {result.end_time.strftime('%H:%M')}",
                audio_file_path="",  # Will be set by caller if needed
                audio_duration_seconds=result.audio_duration,
                cortinillas_found=result.total_cortinillas,
                cortinillas={phrase: [{"start_time": occ.start_time, "end_time": occ.end_time, "start_seconds": occ.start_seconds, "end_seconds": occ.end_seconds, "confidence": occ.confidence} for occ in occurrences] for phrase, occurrences in result.cortinillas_details.items()},
                processing_time_seconds=0.0,
                success=True,
                error_message=None,
                overlap_filtered=result.overlap_filtered,
                overlap_duration=result.overlap_duration
            )
            
            # Add new execution
            accumulated.executions.append(execution)
            accumulated.total_executions += 1
            accumulated.total_cortinillas_found += result.total_cortinillas
            accumulated.last_execution = result.timestamp.isoformat()
            
            # Convert to dict for JSON serialization
            data = self._accumulated_results_to_dict(accumulated)
            
            # Add success log
            self.add_log_entry("INFO", f"JSON report updated successfully", f"Channel: {result.channel}")
            
            # Save to JSON file with atomic write
            self._atomic_json_write(json_path, data)
                
            logger.info(f"Updated JSON report for {result.channel}: {json_path}")
            
        except Exception as e:
            # Add error log
            self.add_log_entry("ERROR", f"Failed to update JSON report: {str(e)}", f"Channel: {result.channel}")
            self.error_handler.handle_error(e, context, critical=True)
            raise ReportGenerationError(f"Failed to update JSON report: {e}") from e
    
    def update_excel_report(self, result: CortinillaDetectionResult) -> None:
        """
        Update the Excel report file with new cortinilla results.
        
        Args:
            result: CortinillaResult to add to the report
        """
        context = create_error_context(
            "update_excel_report",
            channel=result.channel,
            timestamp=result.timestamp
        )
        
        try:
            excel_path = self.data_dir / f"{result.channel.lower()}_results.xlsx"
            
            # Load existing results from JSON (which should already include the new execution)
            accumulated = safe_execute(
                self.load_existing_results,
                result.channel,
                default_return=None,
                error_handler=self.error_handler,
                context=f"{context} | load_existing"
            )
            
            if accumulated is None:
                # This shouldn't happen if JSON was updated first, but handle it gracefully
                logger.warning(f"No existing results found for {result.channel}, creating new Excel report")
                accumulated = AccumulatedResults(
                    channel_name=result.channel,
                    total_executions=1,
                    total_cortinillas_found=result.total_cortinillas,
                    last_execution=result.timestamp.isoformat(),
                    executions=[ProcessingExecution(
                        timestamp=result.timestamp.isoformat(),
                        time_range=f"{result.start_time.strftime('%H:%M')} - {result.end_time.strftime('%H:%M')}",
                        audio_file_path="",
                        audio_duration_seconds=result.audio_duration,
                        cortinillas_found=result.total_cortinillas,
                        cortinillas={phrase: [{"start_time": occ.start_time, "end_time": occ.end_time, "start_seconds": occ.start_seconds, "end_seconds": occ.end_seconds, "confidence": occ.confidence} for occ in occurrences] for phrase, occurrences in result.cortinillas_details.items()},
                        processing_time_seconds=0.0,
                        success=True,
                        error_message=None,
                        overlap_filtered=result.overlap_filtered,
                        overlap_duration=result.overlap_duration
                    )]
                )
            
            # Add success log
            self.add_log_entry("INFO", f"Excel report updated successfully", f"Channel: {result.channel}")
            
            # Create Excel workbook with multiple sheets using the data from JSON
            safe_execute(
                self._create_excel_workbook,
                accumulated,
                excel_path,
                error_handler=self.error_handler,
                context=f"{context} | create_workbook"
            )
            
            logger.info(f"Updated Excel report for {result.channel}: {excel_path}")
            
        except Exception as e:
            # Add error log
            self.add_log_entry("ERROR", f"Failed to update Excel report: {str(e)}", f"Channel: {result.channel}")
            self.error_handler.handle_error(e, context, critical=True)
            raise ReportGenerationError(f"Failed to update Excel report: {e}") from e
    
    def load_existing_results(self, channel: str) -> Optional[AccumulatedResults]:
        """
        Load existing results from JSON file.
        
        Args:
            channel: Channel name to load results for
            
        Returns:
            AccumulatedResults if file exists, None otherwise
        """
        try:
            json_path = self.data_dir / f"{channel.lower()}_results.json"
            
            if not json_path.exists():
                return None
                
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
            return self._dict_to_accumulated_results(data)
            
        except Exception as e:
            logger.error(f"Error loading existing results for {channel}: {e}")
            return None
    
    def get_channel_summary(self, channel: str) -> Dict:
        """
        Get summary statistics for a channel.
        
        Args:
            channel: Channel name
            
        Returns:
            Dictionary with summary statistics
        """
        accumulated = self.load_existing_results(channel)
        if not accumulated:
            return {
                "channel": channel,
                "total_hours": 0,
                "total_cortinillas": 0,
                "cortinillas_by_type": {},
                "last_processed": None
            }
        
        total_cortinillas = accumulated.total_cortinillas_found
        cortinillas_by_type = {}
        
        for execution in accumulated.executions:
            for cortinilla_type, occurrences in execution.cortinillas.items():
                count = len(occurrences)
                cortinillas_by_type[cortinilla_type] = cortinillas_by_type.get(cortinilla_type, 0) + count
        
        return {
            "channel": channel,
            "total_executions": accumulated.total_executions,
            "total_cortinillas": total_cortinillas,
            "cortinillas_by_type": cortinillas_by_type,
            "last_processed": accumulated.last_execution
        }
    
    def _create_excel_workbook(self, accumulated: AccumulatedResults, excel_path: Path) -> None:
        """Create Excel workbook with formatted sheets."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        self._create_summary_sheet(wb, accumulated)
        
        # Create detailed results sheet
        self._create_details_sheet(wb, accumulated)
        
        # Create cortinillas breakdown sheet
        self._create_breakdown_sheet(wb, accumulated)
        
        # Create logs and errors sheet
        self._create_logs_sheet(wb, accumulated)
        
        # Save workbook
        wb.save(excel_path)
    
    def _create_summary_sheet(self, wb: Workbook, accumulated: AccumulatedResults) -> None:
        """Create summary sheet with channel statistics."""
        ws = wb.create_sheet("Resumen")
        
        # Headers
        headers = ["Métrica", "Valor"]
        ws.append(headers)
        
        # Apply professional header formatting
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Summary data
        total_cortinillas = accumulated.total_cortinillas_found
        avg_cortinillas = total_cortinillas / accumulated.total_executions if accumulated.total_executions else 0
        
        summary_data = [
            ["Total Ejecuciones", accumulated.total_executions],
            ["Total Cortinillas Detectadas", total_cortinillas],
            ["Promedio Cortinillas por Ejecución", f"{avg_cortinillas:.2f}"],
            ["Última Ejecución", self._format_datetime(accumulated.last_execution) if accumulated.last_execution else "N/A"]
        ]
        
        # Add data rows with formatting
        for row_idx, row_data in enumerate(summary_data, start=2):
            ws.append(row_data)
            
            # Apply alternating row colors
            if row_idx % 2 == 0:
                fill_color = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for col_idx in range(1, 3):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
                
                # Bold font for metric names
                if col_idx == 1:
                    cell.font = Font(bold=True, size=11)
                else:
                    cell.font = Font(size=11)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # Add borders
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
    
    def _create_details_sheet(self, wb: Workbook, accumulated: AccumulatedResults) -> None:
        """Create detailed results sheet."""
        ws = wb.create_sheet("Detalles por Hora")
        
        # Add cortinilla types as columns first
        all_cortinillas = set()
        for execution in accumulated.executions:
            all_cortinillas.update(execution.cortinillas.keys())
        
        cortinilla_headers = sorted(list(all_cortinillas))
        
        # Headers - reorganized with overlap info at the end
        headers = [
            "Rango de Tiempo",
            "Duración (min)",
            "Total Cortinillas"
        ]
        headers.extend(cortinilla_headers)
        headers.extend([
            "Solapamiento Detectado",
            "Duración Solapamiento (min)"
        ])
        
        ws.append(headers)
        
        # Apply professional header formatting
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_idx, execution in enumerate(accumulated.executions, start=2):
            row = [
                execution.time_range,
                int(round(execution.audio_duration_seconds / 60)),
                execution.cortinillas_found
            ]
            
            # Add cortinilla counts
            for cortinilla in cortinilla_headers:
                count = len(execution.cortinillas.get(cortinilla, []))
                row.append(count)
            
            # Add overlap information at the end
            overlap_detected = "Sí" if getattr(execution, 'overlap_filtered', False) else "No"
            overlap_duration = f"{getattr(execution, 'overlap_duration', 0.0) / 60:.1f}" if getattr(execution, 'overlap_duration', 0.0) > 0 else "0.0"
            
            row.extend([overlap_detected, overlap_duration])
            
            ws.append(row)
            
            # Apply alternating row colors for better readability
            if row_idx % 2 == 0:
                fill_color = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Color code cortinilla counts
                if col_idx > 3 and col_idx <= 3 + len(cortinilla_headers):
                    if cell.value and cell.value > 0:
                        cell.font = Font(bold=True, color="2E7D32")  # Green for detected cortinillas
                
                # Color code overlap detection
                if col_idx == len(row) - 1:  # Overlap detected column
                    if cell.value == "Sí":
                        cell.font = Font(bold=True, color="D32F2F")  # Red for overlap detected
                    else:
                        cell.font = Font(color="757575")  # Gray for no overlap
        
        # Auto-adjust column widths with better sizing
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 3, 12), 25)  # Min 12, max 25
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders to all cells
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
    
    def _create_breakdown_sheet(self, wb: Workbook, accumulated: AccumulatedResults) -> None:
        """Create cortinillas breakdown sheet."""
        ws = wb.create_sheet("Desglose Cortinillas")
        
        # Calculate totals by cortinilla type
        cortinillas_totals = {}
        for execution in accumulated.executions:
            for cortinilla_type, occurrences in execution.cortinillas.items():
                count = len(occurrences)
                cortinillas_totals[cortinilla_type] = cortinillas_totals.get(cortinilla_type, 0) + count
        
        # Headers
        headers = ["Tipo de Cortinilla", "Total Detecciones", "Promedio por Ejecución"]
        ws.append(headers)
        
        # Apply professional header formatting
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        total_executions = len(accumulated.executions)
        sorted_cortinillas = sorted(cortinillas_totals.items(), key=lambda x: x[1], reverse=True)  # Sort by count descending
        
        for row_idx, (cortinilla_type, total_count) in enumerate(sorted_cortinillas, start=2):
            avg_per_execution = total_count / total_executions if total_executions > 0 else 0
            row = [
                cortinilla_type,
                total_count,
                f"{avg_per_execution:.2f}"
            ]
            ws.append(row)
            
            # Apply alternating row colors
            if row_idx % 2 == 0:
                fill_color = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for col_idx in range(1, 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
                
                # Color code based on detection count
                if col_idx == 2:  # Total detections column
                    if total_count > 5:
                        cell.font = Font(bold=True, color="2E7D32", size=11)  # Green for high counts
                    elif total_count > 0:
                        cell.font = Font(bold=True, color="F57C00", size=11)  # Orange for medium counts
                    else:
                        cell.font = Font(color="757575", size=11)  # Gray for zero counts
                else:
                    cell.font = Font(size=11)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        
        # Add borders
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
    
    def _create_logs_sheet(self, wb: Workbook, accumulated: AccumulatedResults) -> None:
        """Create logs and errors sheet."""
        ws = wb.create_sheet("Logs y Errores")
        
        # Headers
        headers = ["Fecha y Hora", "Nivel", "Mensaje", "Contexto"]
        ws.append(headers)
        
        # Apply professional header formatting
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add log entries
        if self.logs_and_errors:
            for row_idx, log_entry in enumerate(self.logs_and_errors[-50:], start=2):  # Last 50 entries
                row = [
                    self._format_datetime(log_entry["timestamp"]),
                    log_entry["level"],
                    log_entry["message"][:80] + "..." if len(log_entry["message"]) > 80 else log_entry["message"],
                    log_entry["context"][:40] + "..." if len(log_entry["context"]) > 40 else log_entry["context"]
                ]
                ws.append(row)
                
                # Apply alternating row colors
                if row_idx % 2 == 0:
                    base_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                else:
                    base_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                # Color code by log level with professional colors
                level_colors = {
                    "ERROR": {"bg": "FFEBEE", "font": "C62828"},      # Light red background, dark red text
                    "CRITICAL": {"bg": "FFCDD2", "font": "B71C1C"},   # Darker red background, darker red text
                    "WARNING": {"bg": "FFF8E1", "font": "E65100"},    # Light amber background, dark orange text
                    "INFO": {"bg": "E8F5E8", "font": "2E7D32"}        # Light green background, dark green text
                }
                
                for col_idx in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if col_idx == 2:  # Level column
                        level = log_entry["level"]
                        if level in level_colors:
                            cell.fill = PatternFill(start_color=level_colors[level]["bg"], 
                                                  end_color=level_colors[level]["bg"], fill_type="solid")
                            cell.font = Font(bold=True, color=level_colors[level]["font"], size=10)
                        else:
                            cell.fill = base_fill
                            cell.font = Font(size=10)
                    else:
                        cell.fill = base_fill
                        cell.font = Font(size=10)
                    
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            # No logs available
            ws.append(["No hay logs disponibles", "", "", ""])
            for col_idx in range(1, 5):
                cell = ws.cell(row=2, column=col_idx)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                cell.font = Font(italic=True, color="757575", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths for better readability
        ws.column_dimensions['A'].width = 18  # Fecha y Hora
        ws.column_dimensions['B'].width = 12  # Nivel
        ws.column_dimensions['C'].width = 50  # Mensaje
        ws.column_dimensions['D'].width = 25  # Contexto
        
        # Set row height for better readability
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        # Add borders
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
   
    def _accumulated_results_to_dict(self, accumulated: AccumulatedResults) -> Dict:
        """Convert AccumulatedResults to dictionary for JSON serialization."""
        return {
            "total_executions": accumulated.total_executions,
            "total_cortinillas_found": accumulated.total_cortinillas_found,
            "last_execution": self._format_datetime(accumulated.last_execution) if accumulated.last_execution else "N/A",
            "executions": [self._execution_to_dict(execution) for execution in accumulated.executions],
            "logs_and_errors": [
                {
                    "timestamp": self._format_datetime(log["timestamp"]),
                    "level": log["level"],
                    "message": log["message"],
                    "context": log["context"]
                }
                for log in self.logs_and_errors[-20:]  # Last 20 log entries
            ]
        }
    
    def _execution_to_dict(self, execution: ProcessingExecution) -> Dict:
        """Convert ProcessingExecution to dictionary for JSON serialization."""
        return {
            "timestamp": self._format_datetime(execution.timestamp),
            "time_range": execution.time_range,
            "audio_duration_minutes": int(round(execution.audio_duration_seconds / 60)),
            "cortinillas_found": execution.cortinillas_found,
            "cortinillas": execution.cortinillas,
            "processing_time_seconds": execution.processing_time_seconds,
            "success": execution.success,
            "error_message": execution.error_message,
            "overlap_detected": getattr(execution, 'overlap_filtered', False),
            "overlap_duration_minutes": round(getattr(execution, 'overlap_duration', 0.0) / 60, 1) if getattr(execution, 'overlap_duration', 0.0) > 0 else 0.0
        }
    
    def _cortinilla_result_to_dict(self, result: CortinillaResult) -> Dict:
        """Convert CortinillaResult to dictionary for JSON serialization."""
        return {
            "channel": result.channel,
            "timestamp": result.timestamp.isoformat(),
            "audio_duration": result.audio_duration,
            "total_cortinillas": result.total_cortinillas,
            "cortinillas_by_type": result.cortinillas_by_type,
            "cortinillas_details": {
                cortinilla_type: [self._occurrence_to_dict(occ) for occ in occurrences]
                for cortinilla_type, occurrences in result.cortinillas_details.items()
            },
            "overlap_filtered": result.overlap_filtered,
            "overlap_duration": result.overlap_duration
        }
    
    def _occurrence_to_dict(self, occurrence: Occurrence) -> Dict:
        """Convert Occurrence to dictionary for JSON serialization."""
        return {
            "start_time": occurrence.start_time,
            "end_time": occurrence.end_time,
            "confidence": occurrence.confidence,
            "text": occurrence.text
        }
    
    def _dict_to_accumulated_results(self, data: Dict) -> AccumulatedResults:
        """Convert dictionary to AccumulatedResults object."""
        executions = []
        for execution_data in data.get("executions", []):
            executions.append(self._dict_to_execution(execution_data))
        
        # Load logs if available
        if "logs_and_errors" in data:
            self.logs_and_errors = data["logs_and_errors"]
        
        # Handle both old and new format
        channel_name = data.get("channel_name", "Unknown")
        
        return AccumulatedResults(
            channel_name=channel_name,
            total_executions=data["total_executions"],
            total_cortinillas_found=data["total_cortinillas_found"],
            last_execution=data.get("last_execution"),
            executions=executions
        )
    
    def _dict_to_execution(self, data: Dict) -> ProcessingExecution:
        """Convert dictionary to ProcessingExecution object."""
        # Handle both old and new format for audio duration
        audio_duration_seconds = data.get("audio_duration_seconds")
        if audio_duration_seconds is None and "audio_duration_minutes" in data:
            audio_duration_seconds = data["audio_duration_minutes"] * 60
        
        # Handle overlap duration
        overlap_duration = data.get("overlap_duration_minutes", 0.0) * 60 if data.get("overlap_duration_minutes") else 0.0
        
        return ProcessingExecution(
            timestamp=data["timestamp"],
            time_range=data["time_range"],
            audio_file_path=data.get("audio_file_path", ""),
            audio_duration_seconds=audio_duration_seconds or 0.0,
            cortinillas_found=data["cortinillas_found"],
            cortinillas=data["cortinillas"],
            processing_time_seconds=data.get("processing_time_seconds", 0.0),
            success=data.get("success", True),
            error_message=data.get("error_message"),
            overlap_filtered=data.get("overlap_detected", False),
            overlap_duration=overlap_duration
        )
    
    def _dict_to_cortinilla_result(self, data: Dict) -> CortinillaResult:
        """Convert dictionary to CortinillaResult object."""
        cortinillas_details = {}
        for cortinilla_type, occurrences_data in data.get("cortinillas_details", {}).items():
            occurrences = [self._dict_to_occurrence(occ_data) for occ_data in occurrences_data]
            cortinillas_details[cortinilla_type] = occurrences
        
        return CortinillaResult(
            channel=data["channel"],
            timestamp=datetime.fromisoformat(data["timestamp"]),
            audio_duration=data["audio_duration"],
            total_cortinillas=data["total_cortinillas"],
            cortinillas_by_type=data["cortinillas_by_type"],
            cortinillas_details=cortinillas_details,
            overlap_filtered=data["overlap_filtered"],
            overlap_duration=data.get("overlap_duration")
        )
    
    def _dict_to_occurrence(self, data: Dict) -> Occurrence:
        """Convert dictionary to Occurrence object."""
        return Occurrence(
            start_time=data["start_time"],
            end_time=data["end_time"],
            confidence=data["confidence"],
            text=data["text"]
        )
    
    def _atomic_json_write(self, file_path: Path, data: dict) -> None:
        """
        Write JSON data atomically to prevent corruption.
        
        Args:
            file_path: Path to write to
            data: Data to write
        """
        temp_path = file_path.with_suffix(file_path.suffix + ".tmp")
        
        try:
            # Write to temporary file first
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            # Atomic rename
            temp_path.replace(file_path)
            
        except Exception as e:
            # Clean up temp file if it exists
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            raise FileOperationError(f"Failed to write JSON file: {e}") from e


def create_sample_report(channel: str, data_dir: str = "data") -> None:
    """
    Create a sample report for testing purposes.
    
    Args:
        channel: Channel name
        data_dir: Directory where reports will be stored
    """
    from datetime import timedelta
    
    generator = ReportGenerator(data_dir)
    
    # Create sample cortinilla results
    sample_results = []
    base_time = datetime.now() - timedelta(hours=3)
    
    for i in range(3):
        timestamp = base_time + timedelta(hours=i)
        
        # Sample occurrences
        occurrences = {
            "buenos días": [
                Occurrence(start_time="00:02:00", end_time="00:02:02", start_seconds=120.5, end_seconds=122.3, confidence=0.95),
                Occurrence(start_time="00:30:00", end_time="00:30:02", start_seconds=1800.2, end_seconds=1802.1, confidence=0.92)
            ],
            "buenas tardes": [
                Occurrence(start_time="00:15:00", end_time="00:15:02", start_seconds=900.1, end_seconds=901.8, confidence=0.88)
            ]
        }
        
        cortinillas_by_type = {
            cortinilla_type: len(occs) for cortinilla_type, occs in occurrences.items()
        }
        
        result = CortinillaResult(
            channel=channel,
            timestamp=timestamp,
            audio_duration=3600.0,  # 1 hour
            total_cortinillas=sum(cortinillas_by_type.values()),
            cortinillas_by_type=cortinillas_by_type,
            cortinillas_details=occurrences,
            overlap_filtered=i > 0,  # First hour not filtered
            overlap_duration=30.0 if i > 0 else None
        )
        
        sample_results.append(result)
    
    # Generate reports for each sample result
    for result in sample_results:
        generator.update_json_report(result)
        generator.update_excel_report(result)
    
    print(f"Sample reports created for channel '{channel}' in '{data_dir}' directory")


if __name__ == "__main__":
    # Create sample reports for testing
    create_sample_report("Canal1")
    create_sample_report("Canal2")